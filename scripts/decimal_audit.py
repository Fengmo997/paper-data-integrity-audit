#!/usr/bin/env python3
"""Audit numeric columns for decimal precision and digit-pattern anomalies."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable


NUMERIC_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?%?$")
LONG_DECIMAL_PLACES = 3
LAST_DIGIT_ENTROPY_THRESHOLD = 2.5
LAST_DIGIT_CHISQ_THRESHOLD = 27.88
FLOAT_ARTIFACT_MIN_DECIMAL_PLACES = 10
FLOAT_ARTIFACT_MAX_CORRECTION = Decimal("1e-12")


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        text = format(value, ".17g").strip()
    else:
        text = str(value).strip()
    if text.lower() in {"", "na", "nan", "none", "null", "inf", "-inf"}:
        return ""
    return text


def numeric_text(text: str) -> str:
    return text.replace(",", "").rstrip("%").strip()


def likely_float_display_artifact(text: str) -> tuple[bool, str]:
    """Detect decimal tails caused by binary-float display, such as 1.20999999999."""
    raw = numeric_text(text)
    if not raw or "." not in raw or "e" in raw.lower():
        return False, text
    fraction = raw.split(".", 1)[1]
    if len(fraction) < FLOAT_ARTIFACT_MIN_DECIMAL_PLACES:
        return False, text
    if "000000" not in fraction and "999999" not in fraction:
        return False, text
    try:
        value = Decimal(raw)
    except InvalidOperation:
        return False, text

    suffix = "%" if text.strip().endswith("%") else ""
    for scale in range(0, 7):
        quantum = Decimal(1).scaleb(-scale)
        try:
            rounded = value.quantize(quantum)
        except InvalidOperation:
            continue
        if abs(value - rounded) <= FLOAT_ARTIFACT_MAX_CORRECTION:
            normalized = format(rounded, "f")
            return True, f"{normalized}{suffix}"
    return False, text


def canonical_numeric_text(text: str) -> str:
    is_artifact, normalized = likely_float_display_artifact(text)
    return normalized if is_artifact else text


def first_excel_format_section(number_format: str | None) -> str:
    if not number_format:
        return ""
    return str(number_format).split(";", 1)[0]


def strip_excel_format_literals(section: str) -> str:
    output: list[str] = []
    in_quote = False
    index = 0
    while index < len(section):
        char = section[index]
        if char == '"':
            in_quote = not in_quote
            index += 1
            continue
        if in_quote:
            index += 1
            continue
        if char in {"\\", "_", "*"}:
            index += 2
            continue
        if char == "[":
            close = section.find("]", index + 1)
            if close != -1:
                index = close + 1
                continue
        output.append(char)
        index += 1
    return "".join(output)


def decimal_format_from_excel_number_format(number_format: str | None) -> tuple[int, int] | None:
    section = strip_excel_format_literals(first_excel_format_section(number_format))
    if not section or "general" in section.lower() or "e" in section.lower():
        return None
    if "." not in section:
        return (0, 0) if any(char in section for char in "0#?") else None
    decimal_part = section.split(".", 1)[1]
    decimal_tokens = [char for char in decimal_part if char in "0#?"]
    if not decimal_tokens and not any(char in section for char in "0#?"):
        return None
    max_places = len(decimal_tokens)
    required_places = sum(1 for char in decimal_tokens if char == "0")
    return max_places, required_places


def percent_scale_from_excel_number_format(number_format: str | None) -> int:
    section = strip_excel_format_literals(first_excel_format_section(number_format))
    return section.count("%")


def format_excel_numeric_cell(value: object, number_format: str | None) -> str:
    decimal_format = decimal_format_from_excel_number_format(number_format)
    if decimal_format is None:
        return clean_cell(value)
    max_places, required_places = decimal_format
    try:
        decimal_value = Decimal(str(value))
    except InvalidOperation:
        return clean_cell(value)
    percent_scale = percent_scale_from_excel_number_format(number_format)
    if percent_scale:
        decimal_value *= Decimal(100) ** percent_scale
    try:
        decimal_value = decimal_value.quantize(Decimal(1).scaleb(-max_places))
    except InvalidOperation:
        return clean_cell(value)
    text = format(decimal_value, f".{max_places}f")
    if max_places > required_places and "." in text:
        whole, fraction = text.split(".", 1)
        fraction = fraction.rstrip("0")
        if len(fraction) < required_places:
            fraction = fraction + ("0" * (required_places - len(fraction)))
        text = f"{whole}.{fraction}" if fraction else whole
    if percent_scale:
        text += "%"
    return text


def is_numeric_like(text: str) -> bool:
    return bool(NUMERIC_RE.match(numeric_text(text) + ("%" if text.endswith("%") else ""))) or bool(
        NUMERIC_RE.match(numeric_text(text))
    )


def decimal_places(text: str) -> int | None:
    raw = numeric_text(text)
    if not raw:
        return None
    if "." in raw and "e" not in raw.lower():
        return len(raw.split(".", 1)[1])
    try:
        value = Decimal(raw)
    except InvalidOperation:
        return None
    exponent = value.as_tuple().exponent
    return max(0, -exponent)


def numeric_value(text: str) -> float | None:
    raw = numeric_text(text)
    try:
        value = float(raw)
    except ValueError:
        return None
    if not math.isfinite(value):
        return None
    return value


def last_digit(text: str) -> str | None:
    digits = re.sub(r"\D", "", numeric_text(text))
    return digits[-1] if digits else None


def decimal_tail(text: str) -> str | None:
    raw = numeric_text(text)
    if "." not in raw or "e" in raw.lower():
        return None
    tail = raw.split(".", 1)[1]
    return tail[-LONG_DECIMAL_PLACES:] if len(tail) >= LONG_DECIMAL_PLACES else None


def shannon_entropy(items: list[str]) -> float:
    if not items:
        return math.nan
    counts = Counter(items)
    total = len(items)
    return -sum((count / total) * math.log2(count / total) for count in counts.values())


def uniform_chisq(items: list[str], categories: list[str]) -> float:
    if not items:
        return math.nan
    counts = Counter(items)
    expected = len(items) / len(categories)
    if expected == 0:
        return math.nan
    return sum((counts.get(category, 0) - expected) ** 2 / expected for category in categories)


def read_csv_like(path: Path, delimiter: str | None) -> dict[str, list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        if delimiter is None:
            sample = handle.read(4096)
            handle.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
            delimiter = dialect.delimiter
        reader = csv.DictReader(handle, delimiter=delimiter)
        rows = [{k: clean_cell(v) for k, v in row.items()} for row in reader]
    return {path.stem: rows}


def unique_headers(header_cells: Iterable[object]) -> list[str]:
    headers: list[str] = []
    counts: Counter[str] = Counter()
    for index, cell in enumerate(header_cells, start=1):
        base = clean_cell(cell) or f"column_{index}"
        if counts[base]:
            header = f"{base}.{counts[base]}"
        else:
            header = base
        counts[base] += 1
        headers.append(header)
    return headers


def read_excel_openpyxl(path: Path, sheet: str | None) -> dict[str, list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Reading XLSX/XLSM files requires openpyxl.") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_names = [sheet] if sheet else workbook.sheetnames
    result: dict[str, list[dict[str, str]]] = {}
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"Sheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows()
        try:
            header_row = next(iterator)
        except StopIteration:
            result[str(sheet_name)] = []
            continue
        headers = unique_headers(cell.value for cell in header_row)
        rows: list[dict[str, str]] = []
        for row in iterator:
            record: dict[str, str] = {}
            for index, header in enumerate(headers):
                cell = row[index] if index < len(row) else None
                if cell is None:
                    record[header] = ""
                elif isinstance(cell.value, (int, float, Decimal)):
                    record[header] = format_excel_numeric_cell(cell.value, cell.number_format)
                else:
                    record[header] = clean_cell(cell.value)
            rows.append(record)
        result[str(sheet_name)] = rows
    workbook.close()
    return result


def read_excel(path: Path, sheet: str | None) -> dict[str, list[dict[str, str]]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_excel_openpyxl(path, sheet)

    try:
        import pandas as pd
    except ImportError as exc:
        raise SystemExit("Reading XLS files requires pandas and an Excel engine.") from exc

    if sheet:
        frames = {sheet: pd.read_excel(path, sheet_name=sheet, dtype=object)}
    else:
        frames = pd.read_excel(path, sheet_name=None, dtype=object)
    result: dict[str, list[dict[str, str]]] = {}
    for sheet_name, frame in frames.items():
        frame = frame.where(pd.notna(frame), "")
        result[str(sheet_name)] = [
            {str(k): clean_cell(v) for k, v in record.items()}
            for record in frame.to_dict(orient="records")
        ]
    return result


def load_tables(path: Path, sheet: str | None, delimiter: str | None) -> dict[str, list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls", ".xlsm"}:
        return read_excel(path, sheet)
    if suffix in {".csv", ".tsv", ".txt"}:
        if suffix == ".tsv" and delimiter is None:
            delimiter = "\t"
        return read_csv_like(path, delimiter)
    raise SystemExit(f"Unsupported input type: {suffix}")


def column_names(rows: Iterable[dict[str, str]]) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for name in row:
            if name not in seen:
                names.append(name)
                seen.add(name)
    return names


def audit_column(sheet: str, column: str, values: list[str], denominator: int | None) -> dict[str, object] | None:
    nonempty = [clean_cell(v) for v in values]
    nonempty = [v for v in nonempty if v]
    numeric_source_values = [v for v in nonempty if is_numeric_like(v)]
    if not numeric_source_values:
        return None
    numeric_fraction = len(numeric_source_values) / max(1, len(nonempty))
    if numeric_fraction < 0.8:
        return None

    numeric_values: list[str] = []
    float_artifact_examples: list[dict[str, str]] = []
    float_artifact_normalized_count = 0
    for value in numeric_source_values:
        is_artifact, normalized = likely_float_display_artifact(value)
        if is_artifact:
            float_artifact_normalized_count += 1
            if len(float_artifact_examples) < 5:
                float_artifact_examples.append({"raw": value, "normalized": normalized})
        numeric_values.append(normalized if is_artifact else value)

    places = [decimal_places(v) for v in numeric_values]
    places = [p for p in places if p is not None]
    digits = [last_digit(v) for v in numeric_values]
    digits = [d for d in digits if d is not None]
    tails = [decimal_tail(v) for v in numeric_values]
    tails = [t for t in tails if t is not None]
    parsed = [numeric_value(v) for v in numeric_values]
    parsed = [v for v in parsed if v is not None]

    value_counts = Counter(numeric_values)
    repeated_values = {k: v for k, v in value_counts.items() if v > 1}
    place_counts = Counter(places)
    digit_counts = Counter(digits)
    tail_counts = Counter(tails)
    repeated_tails = {k: v for k, v in tail_counts.items() if v >= 3}

    long_decimal_count_ge_3 = sum(1 for p in places if p >= LONG_DECIMAL_PLACES)
    long_decimal_count_ge_6 = sum(1 for p in places if p >= 6)
    integer_as_decimal_count = 0
    for text, value in zip(numeric_values, parsed):
        if "." in numeric_text(text) and float(value).is_integer():
            integer_as_decimal_count += 1

    zero_five_count = sum(1 for d in digits if d in {"0", "5"})
    zero_five_rate = zero_five_count / len(digits) if digits else None
    last_digit_entropy = shannon_entropy(digits)
    last_digit_chisq = uniform_chisq(digits, list("0123456789"))

    percent_increment_issue = ""
    if denominator and denominator > 0:
        step = 100.0 / denominator
        bad = []
        for value in parsed:
            if 0 <= value <= 100:
                nearest = round(value / step) * step
                if abs(value - nearest) > 1e-6:
                    bad.append(value)
        if bad:
            percent_increment_issue = f"{len(bad)} values incompatible with denominator {denominator}"

    issues: list[str] = []
    if len(place_counts) > 2:
        issues.append("mixed_decimal_precision")
    if long_decimal_count_ge_3:
        issues.append("long_decimal_precision")
    if integer_as_decimal_count:
        issues.append("integer_reported_as_decimal")
    if zero_five_rate is not None and len(digits) >= 20 and zero_five_rate >= 0.45:
        issues.append("last_digit_0_or_5_enrichment")
    if len(digits) >= 20 and math.isfinite(last_digit_entropy) and last_digit_entropy < LAST_DIGIT_ENTROPY_THRESHOLD:
        issues.append("low_last_digit_entropy")
    if len(digits) >= 20 and math.isfinite(last_digit_chisq) and last_digit_chisq >= LAST_DIGIT_CHISQ_THRESHOLD:
        issues.append("last_digit_nonuniform_p_lt_0_001_reference")
    if repeated_tails:
        issues.append("repeated_decimal_tails")
    if percent_increment_issue:
        issues.append("percentage_denominator_incompatibility")

    return {
        "sheet": sheet,
        "column": column,
        "nonempty_n": len(nonempty),
        "numeric_n": len(numeric_values),
        "float_artifact_normalized_count": float_artifact_normalized_count,
        "float_artifact_examples": json.dumps(float_artifact_examples, ensure_ascii=False),
        "unique_numeric_values": len(value_counts),
        "duplicate_value_count": sum(v - 1 for v in value_counts.values() if v > 1),
        "repeated_values": json.dumps(repeated_values, sort_keys=True),
        "decimal_place_counts": json.dumps(dict(sorted(place_counts.items())), sort_keys=True),
        "max_decimal_places": max(places) if places else "",
        "long_decimal_count_ge_3": long_decimal_count_ge_3,
        "long_decimal_count_ge_6": long_decimal_count_ge_6,
        "integer_as_decimal_count": integer_as_decimal_count,
        "last_digit_counts": json.dumps(dict(sorted(digit_counts.items())), sort_keys=True),
        "last_digit_entropy_log2": round(last_digit_entropy, 9) if math.isfinite(last_digit_entropy) else "",
        "last_digit_chisq_uniform_df9": round(last_digit_chisq, 9) if math.isfinite(last_digit_chisq) else "",
        "last_digit_0_or_5_rate": round(zero_five_rate, 6) if zero_five_rate is not None else "",
        "repeated_decimal_tails": json.dumps(repeated_tails, sort_keys=True),
        "percentage_denominator_issue": percent_increment_issue,
        "issue_flags": ";".join(issues),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="CSV/TSV/TXT/XLS/XLSX table")
    parser.add_argument("--sheet", help="Excel sheet name; default audits all sheets")
    parser.add_argument("--delimiter", help="CSV delimiter; default detects comma/tab/semicolon")
    parser.add_argument("--denominator", type=int, help="Known denominator for percentage compatibility checks")
    parser.add_argument("--out", default="decimal_audit.csv", help="Output CSV path")
    args = parser.parse_args()

    tables = load_tables(Path(args.input), args.sheet, args.delimiter)
    records: list[dict[str, object]] = []
    for sheet, rows in tables.items():
        for column in column_names(rows):
            values = [row.get(column, "") for row in rows]
            record = audit_column(sheet, column, values, args.denominator)
            if record:
                records.append(record)

    fieldnames = [
        "sheet",
        "column",
        "nonempty_n",
        "numeric_n",
        "float_artifact_normalized_count",
        "float_artifact_examples",
        "unique_numeric_values",
        "duplicate_value_count",
        "repeated_values",
        "decimal_place_counts",
        "max_decimal_places",
        "long_decimal_count_ge_3",
        "long_decimal_count_ge_6",
        "integer_as_decimal_count",
        "last_digit_counts",
        "last_digit_entropy_log2",
        "last_digit_chisq_uniform_df9",
        "last_digit_0_or_5_rate",
        "repeated_decimal_tails",
        "percentage_denominator_issue",
        "issue_flags",
    ]
    with Path(args.out).open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    print(f"Wrote {len(records)} audited numeric columns to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
